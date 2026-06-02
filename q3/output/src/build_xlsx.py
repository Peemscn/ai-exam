# -*- coding: utf-8 -*-
"""รวม data เป็น xlsx 1 ไฟล์ 3 sheets (Raw / Clean / Scoring)
   → deliverables/restaurant_data.xlsx (ใส่ git แทน Google Sheets link — เนื้อหาเดียวกัน)
"""
import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
D="q3/output/data"
raw=json.load(open(f"{D}/raw_maps.json",encoding="utf-8"))
clean=json.load(open(f"{D}/clean.json",encoding="utf-8"))
scored=json.load(open(f"{D}/scored.json",encoding="utf-8"))
SUB=["rating_review","group_suitability","price_suitability","travel_convenience","data_completeness","uniqueness"]

HF=PatternFill("solid",fgColor="15324E"); HFONT=Font(bold=True,color="FFFFFF",size=10)
THIN=Side(style="thin",color="D9D9D9"); BD=Border(*[THIN]*4)
def sheet(ws,headers,rowsdata,widths):
    ws.append(headers)
    for j,_ in enumerate(headers,1):
        c=ws.cell(1,j); c.fill=HF; c.font=HFONT; c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width=widths[j-1]
    for r in rowsdata:
        ws.append(r)
        for j in range(1,len(headers)+1):
            cc=ws.cell(ws.max_row,j); cc.border=BD; cc.font=Font(size=9)
            cc.alignment=Alignment(vertical="top",wrap_text=isinstance(r[j-1],str) and len(str(r[j-1]))>30)
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(headers))}{ws.max_row}"

wb=openpyxl.Workbook()
# 1. Raw (Google Maps) — หลักฐานดิบก่อน clean
ws=wb.active; ws.title="Raw Data (Maps)"
sheet(ws,["name","rating","reviews","price","category","url"],
      [[r.get("name"),r.get("rating"),r.get("reviews"),r.get("price"),r.get("category"),r.get("url")] for r in raw],
      [34,8,9,13,18,60])
# 2. Clean — หลัง normalize
ws2=wb.create_sheet("Clean Data")
sheet(ws2,["name","area","category","rating","reviews","price_text","price_min","price_max","address","distance_m","hours","website","matched_osm"],
      [[c["name"],c["area"],c["category"],c["rating"],c["reviews"],c["price_text"],c.get("price_min"),c.get("price_max"),c.get("address"),c.get("distance_m"),c.get("hours"),c.get("website"),c["matched_osm"]] for c in clean],
      [34,8,18,7,8,13,9,9,40,9,16,30,11])
# 3. Scoring — คะแนน 6 หมวด + รวม
ws3=wb.create_sheet("Scoring")
sheet(ws3,["rank","name","category","rating","reviews","price","distance_m","RatingReview/25","Group/20","Price/15","Travel/15","DataComplete/15","Uniqueness/10","TOTAL/100"],
      [[r["rank"],r["name"],r["category"],r["rating"],r["reviews"],r["price_text"],r["distance_m"],*[r["scores"][k] for k in SUB],r["total"]] for r in scored],
      [6,34,18,7,8,13,9,14,9,9,9,14,12,11])
wb.save("q3/output/deliverables/restaurant_data.xlsx")
print(f"saved restaurant_data.xlsx | Raw {len(raw)} · Clean {len(clean)} · Scoring {len(scored)} แถว (3 sheets)")

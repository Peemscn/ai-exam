# -*- coding: utf-8 -*-
"""ประกอบ feedback_raw.json + classified.jsonl -> feedback_analysis.xlsx
   3 sheets: Note (อธิบายไฟล์), Classified Data (300 rows, color-coded, filter), Data Dictionary
"""
import json, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

raw={r["feedback_id"]:r for r in json.load(open("q1/output/data/feedback_raw.json",encoding="utf-8"))}
cls={}
for line in open("q1/output/data/classified.jsonl",encoding="utf-8"):
    r=json.loads(line); cls[r["feedback_id"]]=r

# join (เรียงตาม feedback_id)
COLS=["feedback_id","date","source","player_id","player_segment","platform","game_version",
      "game_area_hint","player_feedback","sentiment","category","priority","ai_summary",
      "suggested_owner","confidence","review_note"]
rows=[]
for fid in sorted(raw, key=lambda x:int(x.split("-")[1])):
    r=dict(raw[fid]); r.update(cls[fid]); rows.append([r.get(c,"") for c in COLS])

wb=openpyxl.Workbook()

# ---- styles ----
HDR_FILL=PatternFill("solid",fgColor="1F3864"); HDR_FONT=Font(bold=True,color="FFFFFF",size=10)
SENT={"Positive":"C6EFCE","Neutral":"FFEB9C","Negative":"FFC7CE"}
PRIO={"High":"FF9999","Medium":"FFD966","Low":"C6EFCE"}
THIN=Side(style="thin",color="D9D9D9"); BORDER=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
WIDTHS={"feedback_id":9,"date":17,"source":17,"player_id":9,"player_segment":15,"platform":13,
        "game_version":11,"game_area_hint":15,"player_feedback":52,"sentiment":10,"category":21,
        "priority":9,"ai_summary":46,"suggested_owner":17,"confidence":11,"review_note":52}

# ===== Sheet 1: Classified Data =====
ws=wb.active; ws.title="Classified Data"
ws.append(COLS)
for j,c in enumerate(COLS,1):
    cell=ws.cell(1,j); cell.fill=HDR_FILL; cell.font=HDR_FONT
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.column_dimensions[get_column_letter(j)].width=WIDTHS[c]
idx={c:i for i,c in enumerate(COLS)}
for r in rows:
    ws.append(r)
    rownum=ws.max_row
    sfill=SENT.get(r[idx["sentiment"]]); pfill=PRIO.get(r[idx["priority"]])
    if sfill: ws.cell(rownum,idx["sentiment"]+1).fill=PatternFill("solid",fgColor=sfill)
    if pfill: ws.cell(rownum,idx["priority"]+1).fill=PatternFill("solid",fgColor=pfill)
    if r[idx["confidence"]]=="Low":      # flag เคสที่ AI ไม่มั่นใจ ให้คนตรวจเห็นง่าย
        ws.cell(rownum,idx["confidence"]+1).font=Font(bold=True,color="C00000")
    for j in range(1,len(COLS)+1):
        cc=ws.cell(rownum,j); cc.border=BORDER
        cc.alignment=Alignment(vertical="top",wrap_text=(COLS[j-1] in ("player_feedback","ai_summary","review_note")))
        if not (COLS[j-1]=="confidence" and r[idx["confidence"]]=="Low"):
            cc.font=Font(size=9)   # confidence-Low cell ตั้ง bold แดงไว้แล้ว ไม่ทับ
ws.freeze_panes="A2"
ws.auto_filter.ref=f"A1:{get_column_letter(len(COLS))}{ws.max_row}"

# ===== Sheet 2: Data Dictionary =====
dd=wb.create_sheet("Data Dictionary")
DDROWS=[
 ("Column","คำอธิบาย","ค่าที่เป็นได้ / ตัวอย่าง","ที่มา"),
 ("feedback_id","รหัส Feedback","FB-001 .. FB-300","Dataset"),
 ("date","วันเวลาที่ได้รับ","2026-05-xx","Dataset"),
 ("source","ช่องทางที่มา","Discord / App Store / Play Store / Facebook / Google Form / Customer Support / Community Post / In-game Survey","Dataset"),
 ("player_id","รหัสผู้เล่น","Pxxxxx","Dataset"),
 ("player_segment","กลุ่มผู้เล่น","Whale / Guild Leader / Mid-Light Spender / F2P / New / Casual / Returning","Dataset"),
 ("platform","แพลตฟอร์ม","iOS / Android / PC Emulator","Dataset"),
 ("game_version","เวอร์ชัน","v1.8.0 - v1.9.1","Dataset"),
 ("game_area_hint","ส่วนของเกม (อาจว่าง)","Gacha Banner / Arena / Event Shop ...","Dataset"),
 ("player_feedback","ข้อความ feedback ดิบ","(ข้อความเต็มรวม filler)","Dataset"),
 ("sentiment","อารมณ์ (จัดจาก core message)","Positive / Neutral / Negative","AI (Claude)"),
 ("category","หมวดหมู่หลัก 9 หมวด + Other","Bug / Reward-Economy / Gameplay-Balance / Event / Gacha-Monetization / UX-UI / Content Request / Positive / Account-Payment","AI (Claude)"),
 ("priority","ความสำคัญในการแก้ (ปรับตาม segment)","High / Medium / Low","AI (Claude)"),
 ("ai_summary","สรุป feedback 1 ประโยค","(ข้อความ)","AI (Claude)"),
 ("suggested_owner","ทีมที่ควรรับผิดชอบ","Game Design / Product / Game Economy / Live Ops / QA / Monetization / Community / Customer Support","AI (Claude)"),
 ("confidence","ความมั่นใจของ AI","High / Medium / Low (Low = ตัวหนาแดง ควรให้คนตรวจ)","AI (Claude)"),
 ("review_note","เหตุผลการจัด + การปรับ priority + เคสกำกวม","(ข้อความ)","AI (Claude)"),
]
for i,row in enumerate(DDROWS,1):
    dd.append(row)
    for j in range(1,5):
        c=dd.cell(i,j); c.border=BORDER; c.alignment=Alignment(vertical="top",wrap_text=True)
        if i==1: c.fill=HDR_FILL; c.font=HDR_FONT
for col,w in zip("ABCD",(18,40,60,14)): dd.column_dimensions[col].width=w
dd.freeze_panes="A2"

# ===== Sheet 3: Note =====
nt=wb.create_sheet("README",0)
NOTES=[
 ("AI Feedback Classifier — Cleaned & Classified Dataset",),
 ("",),
 ("ไฟล์นี้: player feedback 300 รายการ ที่ผ่านการ classify ด้วย AI (Claude Opus 4.8)",),
 ("",),
 ("Pipeline:",),
 ("  1) Data Prep  — ตรวจ schema/null/duplicate, แยก filler ออกจาก core message",),
 ("  2) AI Classify — Claude อ่านทั้ง 300 รายการ จับเป็น 65 canonical intents",),
 ("                   ตีความ sentiment/category/priority/owner ตามความหมาย",),
 ("                   แล้ว encode เป็น deterministic rules (consistent + reproducible)",),
 ("  3) Priority weighting — ปรับตาม player_segment:",),
 ("        Whale / Guild Leader บ่นเรื่องรายได้ (Gacha/Economy) -> +1 priority",),
 ("        New / Returning เจอปัญหา onboarding/retention      -> +1 priority",),
 ("        Account / Payment (ล็อกอิน/จ่ายเงินถูกบล็อก)        -> High เสมอ",),
 ("",),
 ("Data quality: 300/300 records, 0 duplicate, 0 empty feedback, 0 unmatched intent",),
 ("Validation   : ดูไฟล์ validation_note.md (สุ่มตรวจ 30 รายการ = 10%)",),
 ("รายละเอียดวิธีคิด/Prompt: prompt_log.md, workflow.md",),
 ("",),
 ("Sheets: [Classified Data] ผลเต็ม 300 แถว  |  [Data Dictionary] คำอธิบายคอลัมน์",),
]
for i,row in enumerate(NOTES,1):
    nt.append(row); c=nt.cell(i,1)
    if i==1: c.font=Font(bold=True,size=14,color="1F3864")
    elif row[0].endswith(":") or row[0].startswith("Data quality") or row[0].startswith("Sheets"): c.font=Font(bold=True,size=10)
    else: c.font=Font(size=10)
nt.column_dimensions["A"].width=95

wb.save("q1/output/deliverables/feedback_analysis.xlsx")
print("saved feedback_analysis.xlsx |", len(rows), "rows |", len(COLS), "cols | 3 sheets")
print("Low-confidence rows flagged:", sum(1 for r in rows if r[idx['confidence']]=="Low"))
